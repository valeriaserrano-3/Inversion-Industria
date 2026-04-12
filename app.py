"""
Inversión Automotriz – App industria
Streamlit app para procesar archivos de inversión y generar análisis.
"""

import streamlit as st
import pandas as pd
import numpy as np
import plotly.graph_objects as go
import io
import calendar
import warnings
from datetime import datetime

warnings.filterwarnings("ignore")

# ─────────────────────────────────────────────────────────────────────────────
# CONFIGURACIÓN DE PÁGINA
# ─────────────────────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="Inversión Automotriz",
    page_icon=None,
    layout="wide",
)

# ─────────────────────────────────────────────────────────────────────────────
# CONTRASEÑA – GATE antes de mostrar cualquier dato
# ─────────────────────────────────────────────────────────────────────────────
def check_password() -> bool:
    """Muestra el gate de contraseña. Devuelve True si ya está autenticado."""
    if st.session_state.get("authenticated"):
        return True

    st.markdown(
        """
        <style>
        .login-box {
            max-width: 380px;
            margin: 80px auto 0 auto;
            padding: 40px 36px 32px;
            background: #fff;
            border-radius: 16px;
            box-shadow: 0 4px 24px rgba(0,0,0,.10);
            text-align: center;
        }
        .login-title { font-size: 1.5rem; font-weight: 700; margin-bottom: 6px; }
        .login-sub   { font-size: .9rem; color: #666; margin-bottom: 28px; }
        </style>
        <div class="login-box">
          <div class="login-title">Inversión Automotriz</div>
          <div class="login-sub">Ingresa la contraseña para continuar</div>
        </div>
        """,
        unsafe_allow_html=True,
    )

    col = st.columns([1, 2, 1])[1]
    with col:
        pwd = st.text_input("Contraseña", type="password", key="_pwd_input", label_visibility="collapsed")
        entered = st.button("Entrar", use_container_width=True)

    if entered or (pwd and pwd != ""):
        correct = st.secrets.get("APP_PASSWORD", "automotriz2026")
        if pwd == correct:
            st.session_state["authenticated"] = True
            st.rerun()
        else:
            st.error("Contraseña incorrecta.")

    return False


if not check_password():
    st.stop()

# ─────────────────────────────────────────────────────────────────────────────
# CSS GLOBAL (IBM Plex Sans + animación carro gris)
# ─────────────────────────────────────────────────────────────────────────────
st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=IBM+Plex+Sans:wght@300;400;500;600;700&family=Poppins:wght@600;700&display=swap');

html, body, [class*="css"] { font-family: 'IBM Plex Sans', sans-serif; }
h1, h2, h3 { font-family: 'Poppins', sans-serif; }

/* Pills de fuente */
.pill {
    display: inline-block;
    padding: 4px 14px;
    border-radius: 20px;
    font-size: .78rem;
    font-weight: 600;
    margin: 2px 3px;
    border: 1.5px solid transparent;
    cursor: default;
}
.pill-online   { background: #e8f4f8; color: #0077b6; border-color: #0077b6; }
.pill-radio    { background: #fff3e0; color: #e65100; border-color: #e65100; }
.pill-tele     { background: #f3e5f5; color: #7b1fa2; border-color: #7b1fa2; }
.pill-magazine { background: #e8f5e9; color: #2e7d32; border-color: #2e7d32; }
.pill-newspaper{ background: #fce4ec; color: #c62828; border-color: #c62828; }

/* Métricas */
.metric-card {
    background: #f8f9fa;
    border-radius: 12px;
    padding: 16px 20px;
    text-align: center;
    border: 1px solid #e9ecef;
}
.metric-label { font-size: .78rem; color: #888; font-weight: 500; margin-bottom: 4px; }
.metric-value { font-size: 1.35rem; font-weight: 700; color: #1a1a2e; }

/* Animación carro gris */
@keyframes drive { 0% { transform: translateX(-120px); } 100% { transform: translateX(100vw); } }
@keyframes wheel { from { transform: rotate(0deg); } to { transform: rotate(360deg); } }
@keyframes road  { from { background-position: 0 0; } to { background-position: -200px 0; } }

.car-container {
    position: relative; width: 100%; height: 110px;
    overflow: hidden; margin: 24px 0;
}
.car-road {
    position: absolute; bottom: 0; left: 0; right: 0; height: 18px;
    background: repeating-linear-gradient(90deg,#bbb 0,#bbb 40px,transparent 40px,transparent 80px);
    animation: road 0.6s linear infinite;
    background-color: #ddd;
}
.car-driving {
    position: absolute; bottom: 14px;
    animation: drive 2.4s linear infinite;
}
.car-body { fill: #888888; }
.car-window { fill: #cce4f0; }
.car-wheel-outer { fill: #444; }
.car-wheel-inner { fill: #bbb; }
.car-wheel-group { animation: wheel 0.5s linear infinite; transform-origin: 50% 50%; }
</style>
""", unsafe_allow_html=True)

CAR_SVG = """
<svg xmlns="http://www.w3.org/2000/svg" width="120" height="60" viewBox="0 0 120 60">
  <rect x="10" y="22" width="100" height="28" rx="6" class="car-body"/>
  <polygon points="28,22 38,8 82,8 92,22" class="car-body"/>
  <polygon points="32,22 40,11 80,11 88,22" class="car-window"/>
  <rect x="0"  y="40" width="12" height="8" rx="2" class="car-body"/>
  <rect x="108" y="40" width="12" height="8" rx="2" class="car-body"/>
  <g class="car-wheel-group" style="transform-origin:30px 50px">
    <circle cx="30" cy="50" r="10" class="car-wheel-outer"/>
    <circle cx="30" cy="50" r="4"  class="car-wheel-inner"/>
  </g>
  <g class="car-wheel-group" style="transform-origin:90px 50px">
    <circle cx="90" cy="50" r="10" class="car-wheel-outer"/>
    <circle cx="90" cy="50" r="4"  class="car-wheel-inner"/>
  </g>
</svg>
"""

def show_car_animation():
    st.markdown(
        f"""<div class="car-container">
              <div class="car-driving">{CAR_SVG}</div>
              <div class="car-road"></div>
            </div>""",
        unsafe_allow_html=True,
    )

# ─────────────────────────────────────────────────────────────────────────────
# LISTAS DE GRUPOS
# ─────────────────────────────────────────────────────────────────────────────
GWM_GRUPOS   = {"NISSAN","CHEVROLET","VW","KIA","BYD","GWM","GEELY","HYUNDAI",
                "CHIREY","GAC","MG","CHANGAN","TOYOTA","OMODA"}
JAC_GRUPOS   = {"JAC","JAC INDUSTRIA","JAC CORP"}
KAVAK_GRUPOS = {"KAVAK","NISSAN","GENERAL MOTORS","HYUNDAI","VOLKSWAGEN","KIA",
                "MITSUBISHI","FORD MOTOR","GEELY","JEEP","INFINITI","PEUGEOT",
                "CHIREY","RENAULT","TOYOTA","MG","BBVA AUTOMARKET","HONDA"}
BAIC_GRUPOS  = {"NISSAN","HYUNDAI","VOLKSWAGEN","KIA","BYD","GEELY","JEEP",
                "TOYOTA","MG","GREAT WALL MOTOR","CHEVROLET","GAC","CHANGAN"}

# ─────────────────────────────────────────────────────────────────────────────
# DATOS HISTÓRICOS 2025 (editables aquí; raramente cambian)
# ─────────────────────────────────────────────────────────────────────────────
# Formato: mes_num → (Online_MXN, Offline_MXN, OOH_MXN)  – en millones × 1_000_000
HIST_2025 = {
     1: (138_000_000, 245_000_000,  91_000_000),
     2: (154_000_000, 295_000_000,  90_000_000),
     3: (123_000_000, 317_000_000, 110_000_000),
     4: (216_000_000, 227_000_000, 142_000_000),
     5: (216_000_000, 226_000_000, 155_000_000),
     6: (355_000_000, 213_000_000, 135_000_000),
     7: (313_000_000, 185_000_000, 144_000_000),
     8: (332_000_000, 182_000_000, 110_000_000),
     9: (287_000_000, 197_000_000, 105_000_000),
    10: (256_000_000, 193_000_000, 138_000_000),
    11: (319_000_000, 237_000_000, 141_000_000),
    12: (360_000_000, 261_000_000, 165_000_000),
}

# ─────────────────────────────────────────────────────────────────────────────
# MULTIPLICADORES TV (ajustables en "Volver a correr")
# ─────────────────────────────────────────────────────────────────────────────
DEFAULT_MULT = {
    "radio":        (0.70, 1.30),
    "tele_abierta": (0.40, 1.30),
    "tele_paga":    (0.60, 1.30),
    "online":       (1.00, 1.30),
    "ooh_mxn_factor": 2.0,
    "ooh_f30_factor": 1.30,
}

_MONTH_ABBR = {m.lower(): i for i, m in enumerate(calendar.month_abbr) if m}
MES_ABBR_EN = {i: m.lower() for i, m in enumerate(calendar.month_abbr) if m}

# ─────────────────────────────────────────────────────────────────────────────
# HELPERS COMUNES
# ─────────────────────────────────────────────────────────────────────────────
def parse_yrmth(val) -> pd.Timestamp:
    """Convierte 'jan2025' / datetime a Timestamp."""
    s = str(val).strip().lower()
    for abbr, num in _MONTH_ABBR.items():
        if s.startswith(abbr):
            try:
                return pd.Timestamp(year=int(s[len(abbr):]), month=num, day=1)
            except ValueError:
                pass
    try:
        return pd.to_datetime(val)
    except Exception:
        return pd.NaT


def fuente_offline_ooh(fuente_str: str) -> str:
    v = str(fuente_str).strip().upper()
    if any(k in v for k in ("TV", "RADIO", "PRENSA", "PERIOD", "CINE", "SPOT")):
        return "Offline"
    return "OOH"


def classify_fuente(medio_str: str, mult: dict) -> str:
    """Determina Fuente (Online/Offline/OOH) a partir del medio."""
    v = str(medio_str).strip().upper()
    if any(k in v for k in ("DIGITAL", "ONLINE", "SEARCH", "SOCIAL", "DISPLAY", "PROGRAMMATIC", "VIDEO")):
        return "Online"
    if any(k in v for k in ("TV", "RADIO", "PRENSA", "PERIOD", "CINE", "MAGAZINE", "NEWSPAPER", "REVISTA")):
        return "Offline"
    if any(k in v for k in ("OOH", "ESPECTACULAR", "MUPIE", "VALLA", "DIGITAL OOH")):
        return "OOH"
    return "Offline"

# ─────────────────────────────────────────────────────────────────────────────
# PROCESADORES POR FUENTE
# ─────────────────────────────────────────────────────────────────────────────

def process_gwm(uploaded_files: list, periods: set, mult: dict) -> pd.DataFrame:
    """Procesa archivos GWM (Offline Naming + Online Resumen)."""
    frames = []
    for uf in uploaded_files:
        try:
            xl = pd.ExcelFile(uf)
            name_up = uf.name.upper()

            # ── ONLINE: hoja "RESUMEN TOTAL" ──────────────────────────────
            if "RESUMEN TOTAL" in xl.sheet_names:
                df = xl.parse("RESUMEN TOTAL")
                # Detectar col de mes (Month / Mes / Año-mes)
                month_col = next((c for c in df.columns if "month" in str(c).lower() or "mes" in str(c).lower()), None)
                model_col  = next((c for c in df.columns if "model" in str(c).lower() or "modelo" in str(c).lower()), None)
                inv_col    = next((c for c in df.columns if "invest" in str(c).lower() or "inver" in str(c).lower()), None)

                if month_col and inv_col:
                    df["_ts"] = df[month_col].apply(parse_yrmth)
                    df = df[df["_ts"].notna()].copy()
                    df["_period"] = df["_ts"].dt.to_period("M")
                    df = df[df["_period"].isin(periods)].copy()
                    df["inv_raw"] = pd.to_numeric(df[inv_col], errors="coerce").fillna(0)
                    df = df[df["inv_raw"] > 0].copy()

                    for _, r in df.iterrows():
                        inv = round(r["inv_raw"] * mult["online"][1], 4)
                        modelo = str(r[model_col]).strip().upper() if model_col else "VARIOS"
                        frames.append({
                            "Fuente": "Online",
                            "Año": float(r["_ts"].year),
                            "Año-mes": r["_ts"].replace(day=1),
                            "Inversión (MXN)": inv,
                            "Inversión F30": inv,
                            "#Grupo": "GWM",
                            "Categoría": "Institucional",
                            "Producto": modelo,
                            "medio": "Digital",
                            "medio2": np.nan,
                            "modelo": modelo,
                            "Inversión original": r["inv_raw"],
                            "No incluir": np.nan,
                        })

            # ── OFFLINE: hoja KAVAK / primera hoja con datos ───────────────
            sheet_offline = None
            for s in xl.sheet_names:
                su = s.upper()
                if "GWM" in su or "NAMING" in su or "NAMING" in name_up:
                    sheet_offline = s
                    break
            if sheet_offline is None and xl.sheet_names:
                sheet_offline = xl.sheet_names[0]

            if sheet_offline and sheet_offline not in (
                "RESUMEN TOTAL",
            ):
                df = xl.parse(sheet_offline)
                if "yrmth" in df.columns:
                    df["_ts"] = df["yrmth"].apply(parse_yrmth)
                    df = df[df["_ts"].notna()].copy()
                    df["_period"] = df["_ts"].dt.to_period("M")
                    df = df[df["_period"].isin(periods)].copy()

                    df["monto"] = pd.to_numeric(df.get("monto", 0), errors="coerce").fillna(0)
                    df["Bonificación Cliente"] = pd.to_numeric(df.get("Bonificación Cliente", 0), errors="coerce").fillna(0)
                    df["inversion"] = df["monto"] + df["Bonificación Cliente"]
                    df = df[df["inversion"] > 0].copy()

                    for _, r in df.iterrows():
                        medio_cat = str(r.get("Medio Catálogo", "")).strip()
                        fuente = "OOH" if "OOH" in medio_cat.upper() else fuente_offline_ooh(str(r.get("FUENTE", "")))
                        modelo = str(r.get("Modelos", "VARIOS")).strip().upper()
                        medio_val = str(r.get("FUENTE", medio_cat)).strip()
                        inv_raw = r["inversion"]
                        if fuente == "OOH":
                            inv_mxn = round(inv_raw * mult["ooh_mxn_factor"], 4)
                            inv_f30 = round(inv_raw * mult["ooh_f30_factor"], 4)
                        else:
                            m1, m2 = _tv_mult(medio_val, mult)
                            inv_mxn = round(inv_raw * m1 * m2, 4)
                            inv_f30 = inv_mxn
                        frames.append({
                            "Fuente": fuente,
                            "Año": float(r["_ts"].year),
                            "Año-mes": r["_ts"].replace(day=1),
                            "Inversión (MXN)": inv_mxn,
                            "Inversión F30": inv_f30,
                            "#Grupo": "GWM",
                            "Categoría": "Institucional",
                            "Producto": modelo,
                            "medio": medio_val,
                            "medio2": np.nan,
                            "modelo": modelo,
                            "Inversión original": inv_raw,
                            "No incluir": np.nan,
                        })
        except Exception as e:
            st.warning(f"GWM – {uf.name}: {e}")

    return pd.DataFrame(frames) if frames else pd.DataFrame()


def process_jac(uploaded_files: list, periods: set, mult: dict) -> pd.DataFrame:
    frames = []
    for uf in uploaded_files:
        try:
            xl = pd.ExcelFile(uf)
            name_up = uf.name.upper()

            # ── ONLINE: "TOTAL FINAL" ───────────────────────────────────────
            if "TOTAL FINAL" in xl.sheet_names:
                df = xl.parse("TOTAL FINAL", header=None)
                # mes del nombre de archivo: "02.FEBRERO.2026.xlsx"
                import re
                m = re.search(r"(\d{2})\.\w+\.(\d{4})", uf.name)
                if m:
                    month_num, year_num = int(m.group(1)), int(m.group(2))
                    ts = pd.Timestamp(year=year_num, month=month_num, day=1)
                    if ts.to_period("M") in periods:
                        # Buscar fila encabezado con "Marca" y columna de inversión
                        header_row = None
                        for i, row in df.iterrows():
                            if any("marca" in str(v).lower() for v in row.values):
                                header_row = i
                                break
                        if header_row is not None:
                            df.columns = df.iloc[header_row]
                            df = df.iloc[header_row + 1:].reset_index(drop=True)
                            marca_col = next((c for c in df.columns if "marca" in str(c).lower()), None)
                            inv_col   = next((c for c in df.columns if "invest" in str(c).lower() or "inver" in str(c).lower()), None)
                            if marca_col and inv_col:
                                for _, r in df.iterrows():
                                    inv_raw = pd.to_numeric(r[inv_col], errors="coerce")
                                    if pd.isna(inv_raw) or inv_raw <= 0:
                                        continue
                                    inv = round(inv_raw * mult["online"][1], 4)
                                    modelo = str(r[marca_col]).strip().upper()
                                    frames.append({
                                        "Fuente": "Online",
                                        "Año": float(ts.year),
                                        "Año-mes": ts,
                                        "Inversión (MXN)": inv,
                                        "Inversión F30": inv,
                                        "#Grupo": "JAC",
                                        "Categoría": "Institucional",
                                        "Producto": modelo,
                                        "medio": "Digital",
                                        "medio2": np.nan,
                                        "modelo": modelo,
                                        "Inversión original": inv_raw,
                                        "No incluir": np.nan,
                                    })

            # ── OFFLINE / OOH: hoja con yrmth ─────────────────────────────
            offline_sheet = next(
                (s for s in xl.sheet_names if s.upper() not in ("TOTAL FINAL",)
                 and "yrmth" in xl.parse(s, nrows=1).columns.tolist()),
                None,
            )
            if offline_sheet:
                df = xl.parse(offline_sheet)
                df["_ts"] = df["yrmth"].apply(parse_yrmth)
                df = df[df["_ts"].notna()].copy()
                df["_period"] = df["_ts"].dt.to_period("M")
                df = df[df["_period"].isin(periods)].copy()
                df["monto"] = pd.to_numeric(df.get("monto", 0), errors="coerce").fillna(0)
                df["Bonificación Cliente"] = pd.to_numeric(df.get("Bonificación Cliente", 0), errors="coerce").fillna(0)
                df["inversion"] = df["monto"] + df["Bonificación Cliente"]
                df = df[df["inversion"] > 0].copy()

                for _, r in df.iterrows():
                    medio_cat = str(r.get("Medio Catálogo", "")).strip()
                    fuente = "OOH" if "OOH" in medio_cat.upper() else fuente_offline_ooh(str(r.get("FUENTE", "")))
                    medio_val = str(r.get("FUENTE", medio_cat)).strip()
                    modelo = str(r.get("Modelos", "VARIOS")).strip().upper()
                    inv_raw = r["inversion"]
                    if fuente == "OOH":
                        inv_mxn = round(inv_raw * mult["ooh_mxn_factor"], 4)
                        inv_f30 = round(inv_raw * mult["ooh_f30_factor"], 4)
                    else:
                        m1, m2 = _tv_mult(medio_val, mult)
                        inv_mxn = round(inv_raw * m1 * m2, 4)
                        inv_f30 = inv_mxn
                    frames.append({
                        "Fuente": fuente,
                        "Año": float(r["_ts"].year),
                        "Año-mes": r["_ts"].replace(day=1),
                        "Inversión (MXN)": inv_mxn,
                        "Inversión F30": inv_f30,
                        "#Grupo": "JAC",
                        "Categoría": "Institucional",
                        "Producto": modelo,
                        "medio": medio_val,
                        "medio2": np.nan,
                        "modelo": modelo,
                        "Inversión original": inv_raw,
                        "No incluir": np.nan,
                    })
        except Exception as e:
            st.warning(f"JAC – {uf.name}: {e}")

    return pd.DataFrame(frames) if frames else pd.DataFrame()


def process_kavak_baic_industria(uploaded_files: list, periods: set, mult: dict) -> pd.DataFrame:
    frames = []
    for uf in uploaded_files:
        try:
            xl = pd.ExcelFile(uf)
            name_up = uf.name.upper()

            # Detectar grupo por nombre
            grupo = "Industria"
            if "KAVAK" in name_up:
                grupo = "KAVAK"
            elif "BAIC" in name_up:
                grupo = "BAIC"

            sheet_name = "KAVAK" if "KAVAK" in xl.sheet_names else xl.sheet_names[0]
            df = xl.parse(sheet_name)

            if "yrmth" not in df.columns:
                st.warning(f"{uf.name}: no se encontró columna 'yrmth'.")
                continue

            df["_ts"] = df["yrmth"].apply(parse_yrmth)
            df = df[df["_ts"].notna()].copy()
            df["_period"] = df["_ts"].dt.to_period("M")
            df = df[df["_period"].isin(periods)].copy()
            df["monto"] = pd.to_numeric(df.get("monto", 0), errors="coerce").fillna(0)
            df["Bonificación Cliente"] = pd.to_numeric(df.get("Bonificación Cliente", 0), errors="coerce").fillna(0)
            df["inversion"] = df["monto"] + df["Bonificación Cliente"]
            df = df[df["inversion"] > 0].copy()

            for _, r in df.iterrows():
                fuente_col = str(r.get("FUENTE", "")).strip()
                fuente = fuente_offline_ooh(fuente_col)
                medio_val = str(r.get("Formato Cátalogo", fuente_col)).strip() or fuente_col
                if fuente == "OOH":
                    medio_val = str(r.get("Formato Cátalogo", "")).strip() or medio_val
                modelo = str(r.get("Modelos", "VARIOS")).strip().upper()
                inv_raw = r["inversion"]
                if fuente == "OOH":
                    inv_mxn = round(inv_raw * mult["ooh_mxn_factor"], 4)
                    inv_f30 = round(inv_raw * mult["ooh_f30_factor"], 4)
                else:
                    m1, m2 = _tv_mult(fuente_col, mult)
                    inv_mxn = round(inv_raw * m1 * m2, 4)
                    inv_f30 = inv_mxn
                frames.append({
                    "Fuente": fuente,
                    "Año": float(r["_ts"].year),
                    "Año-mes": r["_ts"].replace(day=1),
                    "Inversión (MXN)": inv_mxn,
                    "Inversión F30": inv_f30,
                    "#Grupo": grupo,
                    "Categoría": "Institucional",
                    "Producto": modelo,
                    "medio": medio_val,
                    "medio2": np.nan,
                    "modelo": modelo,
                    "Inversión original": inv_raw,
                    "No incluir": np.nan,
                })
        except Exception as e:
            st.warning(f"KAVAK/BAIC – {uf.name}: {e}")

    return pd.DataFrame(frames) if frames else pd.DataFrame()


def process_ooh(uploaded_files: list, periods: set, mult: dict) -> pd.DataFrame:
    """Procesa archivos OOH (formato A/B CSV o XLSX)."""
    frames = []
    for uf in uploaded_files:
        try:
            name_up = uf.name.upper()
            if uf.name.endswith(".csv"):
                df = pd.read_csv(uf, encoding="utf-8-sig")
            else:
                xl = pd.ExcelFile(uf)
                # Probar hoja con datos
                df = xl.parse(xl.sheet_names[0], header=None)
                # Buscar fila encabezado
                header_row = 0
                for i, row in df.iterrows():
                    vals = [str(v).lower() for v in row.values if pd.notna(v)]
                    if any("costo" in v or "fecha" in v or "mes" in v or "marca" in v for v in vals):
                        header_row = i
                        break
                df.columns = df.iloc[header_row]
                df = df.iloc[header_row + 1:].reset_index(drop=True)

            # Detectar cols
            fecha_col = next((c for c in df.columns if "fecha" in str(c).lower() or "mes" in str(c).lower() or "yrmth" in str(c).lower()), None)
            costo_col = next((c for c in df.columns if "costo" in str(c).lower()), None)
            marca_col = next((c for c in df.columns if "marca" in str(c).lower() or "anunciante" in str(c).lower()), None)
            medio_col = next((c for c in df.columns if "formato" in str(c).lower() or "medio" in str(c).lower()), None)
            modelo_col= next((c for c in df.columns if "model" in str(c).lower()), None)

            if not fecha_col or not costo_col:
                st.warning(f"OOH – {uf.name}: no se reconoció estructura (falta fecha o costo).")
                continue

            df["_ts"] = df[fecha_col].apply(parse_yrmth)
            df = df[df["_ts"].notna()].copy()
            df["_period"] = df["_ts"].dt.to_period("M")
            df = df[df["_period"].isin(periods)].copy()
            df["costo"] = pd.to_numeric(df[costo_col], errors="coerce").fillna(0)
            df = df[df["costo"] > 0].copy()

            for _, r in df.iterrows():
                costo = r["costo"]
                inv_mxn = round(costo * mult["ooh_mxn_factor"], 4)
                inv_f30 = round(costo * mult["ooh_f30_factor"], 4)
                marca = str(r[marca_col]).strip().upper() if marca_col else "VARIOS"
                medio_val = str(r[medio_col]).strip() if medio_col else "OOH"
                modelo = str(r[modelo_col]).strip().upper() if modelo_col else marca
                frames.append({
                    "Fuente": "OOH",
                    "Año": float(r["_ts"].year),
                    "Año-mes": r["_ts"].replace(day=1),
                    "Inversión (MXN)": inv_mxn,
                    "Inversión F30": inv_f30,
                    "#Grupo": marca,
                    "Categoría": "Institucional",
                    "Producto": modelo,
                    "medio": medio_val,
                    "medio2": np.nan,
                    "modelo": modelo,
                    "Inversión original": costo,
                    "No incluir": np.nan,
                })
        except Exception as e:
            st.warning(f"OOH – {uf.name}: {e}")

    return pd.DataFrame(frames) if frames else pd.DataFrame()


def process_admetricks(uploaded_files: list, periods: set, mult: dict) -> pd.DataFrame:
    frames = []
    BRAND_MAP = {
        "NISSAN": "NISSAN", "CHEVROLET": "CHEVROLET", "VW": "VW",
        "VOLKSWAGEN": "VW", "KIA": "KIA", "BYD": "BYD", "GWM": "GWM",
        "GEELY": "GEELY", "HYUNDAI": "HYUNDAI", "CHIREY": "CHIREY",
        "GAC": "GAC", "MG": "MG", "CHANGAN": "CHANGAN", "TOYOTA": "TOYOTA",
        "OMODA": "OMODA", "JAC": "JAC", "KAVAK": "KAVAK", "BAIC": "BAIC",
        "FORD": "FORD MOTOR", "JEEP": "JEEP", "HONDA": "HONDA",
        "RENAULT": "RENAULT", "PEUGEOT": "PEUGEOT", "BMW": "BMW",
        "VOLVO": "VOLVO", "LEXUS": "LEXUS",
    }
    for uf in uploaded_files:
        try:
            df = pd.read_excel(uf, sheet_name="Datos 1")
            date_col  = next((c for c in df.columns if "fecha" in str(c).lower() or "date" in str(c).lower()), None)
            inv_col   = next((c for c in df.columns if "inver" in str(c).lower() or "invest" in str(c).lower()), None)
            anun_col  = next((c for c in df.columns if "anunci" in str(c).lower() or "brand" in str(c).lower()), None)
            medio_col = next((c for c in df.columns if "hosped" in str(c).lower() or "site" in str(c).lower() or "medio" in str(c).lower()), None)

            if not date_col or not inv_col:
                st.warning(f"Admetricks – {uf.name}: estructura no reconocida.")
                continue

            df["_ts"] = pd.to_datetime(df[date_col], errors="coerce").dt.to_period("M").dt.to_timestamp()
            df = df[df["_ts"].notna()].copy()
            df["_period"] = df["_ts"].dt.to_period("M")
            df = df[df["_period"].isin(periods)].copy()
            df["inv_raw"] = pd.to_numeric(df[inv_col], errors="coerce").fillna(0)
            df = df[df["inv_raw"] > 0].copy()

            for _, r in df.iterrows():
                inv_raw = r["inv_raw"]
                inv = round(inv_raw * mult["online"][1], 4)
                anun = str(r[anun_col]).strip().upper() if anun_col else "VARIOS"
                grupo = next((v for k, v in BRAND_MAP.items() if k in anun), anun)
                medio_val = str(r[medio_col]).strip() if medio_col else "Digital"
                ts = r["_ts"]
                frames.append({
                    "Fuente": "Online",
                    "Año": float(ts.year),
                    "Año-mes": ts,
                    "Inversión (MXN)": inv,
                    "Inversión F30": inv,
                    "#Grupo": grupo,
                    "Categoría": "Institucional",
                    "Producto": grupo,
                    "medio": medio_val,
                    "medio2": np.nan,
                    "modelo": grupo,
                    "Inversión original": inv_raw,
                    "No incluir": np.nan,
                })
        except Exception as e:
            st.warning(f"Admetricks – {uf.name}: {e}")

    return pd.DataFrame(frames) if frames else pd.DataFrame()


def process_auditsa(uploaded_files: list, periods: set, mult: dict) -> pd.DataFrame:
    frames = []
    for uf in uploaded_files:
        try:
            xl = pd.ExcelFile(uf)
            df = xl.parse(xl.sheet_names[0])
            # Detectar columnas clave
            fecha_col = next((c for c in df.columns if "fecha" in str(c).lower() or "mes" in str(c).lower()), None)
            inv_col   = next((c for c in df.columns if "inver" in str(c).lower() or "monto" in str(c).lower()), None)
            medio_col = next((c for c in df.columns if "medio" in str(c).lower()), None)
            marca_col = next((c for c in df.columns if "marca" in str(c).lower() or "anunciante" in str(c).lower()), None)

            if not fecha_col or not inv_col:
                st.warning(f"Auditsa – {uf.name}: estructura no reconocida.")
                continue

            df["_ts"] = df[fecha_col].apply(parse_yrmth)
            df = df[df["_ts"].notna()].copy()
            df["_period"] = df["_ts"].dt.to_period("M")
            df = df[df["_period"].isin(periods)].copy()
            df["inv_raw"] = pd.to_numeric(df[inv_col], errors="coerce").fillna(0)
            df = df[df["inv_raw"] > 0].copy()

            for _, r in df.iterrows():
                inv_raw = r["inv_raw"]
                medio_val = str(r[medio_col]).strip().upper() if medio_col else ""
                if "RADIO" in medio_val:
                    m1, m2 = mult["radio"]
                    fuente = "Offline"
                elif any(k in medio_val for k in ("MAGAZINE", "REVISTA", "NEWSPAPER", "PRENSA", "PERIOD")):
                    m1, m2 = 1.0, mult["online"][1]
                    fuente = "Offline"
                else:
                    m1, m2 = 1.0, mult["online"][1]
                    fuente = "Offline"
                inv = round(inv_raw * m1 * m2, 4)
                marca = str(r[marca_col]).strip().upper() if marca_col else "VARIOS"
                ts = r["_ts"].replace(day=1)
                frames.append({
                    "Fuente": fuente,
                    "Año": float(ts.year),
                    "Año-mes": ts,
                    "Inversión (MXN)": inv,
                    "Inversión F30": inv,
                    "#Grupo": marca,
                    "Categoría": "Institucional",
                    "Producto": marca,
                    "medio": str(r[medio_col]).strip() if medio_col else "Offline",
                    "medio2": np.nan,
                    "modelo": marca,
                    "Inversión original": inv_raw,
                    "No incluir": np.nan,
                })
        except Exception as e:
            st.warning(f"Auditsa – {uf.name}: {e}")

    return pd.DataFrame(frames) if frames else pd.DataFrame()


def _tv_mult(medio_str: str, mult: dict) -> tuple:
    """Devuelve (m1, m2) según el tipo de TV/Radio/medio."""
    v = str(medio_str).strip().upper()
    if "RADIO" in v:
        return mult["radio"]
    if "ABIERTA" in v or "OPEN" in v:
        return mult["tele_abierta"]
    if "PAGA" in v or "PAGADA" in v or "CABLE" in v or "LOCAL" in v or "REGIONAL" in v:
        return mult["tele_paga"]
    return mult["online"]  # 1.0, 1.3

# ─────────────────────────────────────────────────────────────────────────────
# MERGE EN MEMORIA
# ─────────────────────────────────────────────────────────────────────────────
LAYOUT_COLS = [
    "Fuente","Año","Año-mes","Inversión (MXN)","Inversión F30",
    "#Grupo","Categoría","Producto","medio","medio2","modelo",
    "Inversión original","No incluir",
]

def merge_with_automotriz(df_new: pd.DataFrame, automotriz_bytes: bytes) -> tuple[pd.DataFrame, int, int]:
    """
    Combina df_new con Investment Raw data del xlsx subido.
    Devuelve (df_combinado, filas_antes, filas_despues).
    """
    df_main = pd.read_excel(io.BytesIO(automotriz_bytes), sheet_name="Investment Raw data")
    rows_before = len(df_main)
    df_main["Año-mes"] = pd.to_datetime(df_main["Año-mes"], errors="coerce")

    df_new_save = df_new.copy()
    for col in LAYOUT_COLS:
        if col not in df_new_save.columns:
            df_new_save[col] = np.nan
    df_new_save = df_new_save[LAYOUT_COLS]

    df_combined = pd.concat([df_main, df_new_save], ignore_index=True)
    return df_combined, rows_before, len(df_combined)


def build_automotriz_xlsx(df_combined: pd.DataFrame, automotriz_bytes: bytes) -> bytes:
    """
    Genera bytes de automotriz.xlsx con Investment Raw data actualizado,
    preservando las demás hojas del archivo original.
    """
    from openpyxl import load_workbook

    # Cargar workbook original para conservar otras hojas
    wb = load_workbook(io.BytesIO(automotriz_bytes))

    # Eliminar hoja que vamos a reemplazar
    if "Investment Raw data" in wb.sheetnames:
        del wb["Investment Raw data"]

    # Crear la hoja con pandas y copiar al workbook
    buf_tmp = io.BytesIO()
    with pd.ExcelWriter(buf_tmp, engine="openpyxl") as wr:
        df_combined.to_excel(wr, sheet_name="Investment Raw data", index=False)
    wb_tmp = load_workbook(buf_tmp)
    ws_new = wb_tmp["Investment Raw data"]

    # Mover la hoja nueva al workbook original (posición 0)
    ws_copy = wb.create_sheet("Investment Raw data", 0)
    for row in ws_new.iter_rows():
        for cell in row:
            ws_copy[cell.coordinate] = cell.value

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()

# ─────────────────────────────────────────────────────────────────────────────
# ANÁLISIS Y GRÁFICAS
# ─────────────────────────────────────────────────────────────────────────────
MONTH_ORDER = [f"{calendar.month_abbr[i]}-{str(y)[2:]}"
               for y in [2025, 2026] for i in range(1, 13)]

def build_chart_data(df_main: pd.DataFrame, year: int) -> pd.DataFrame:
    """Construye DataFrame mensual Online/Offline/OOH para el año dado."""
    df = df_main.copy()
    df["_ts"] = pd.to_datetime(df["Año-mes"], errors="coerce")
    df = df[df["_ts"].dt.year == year].copy()

    rows = []
    for m in range(1, 13):
        sub = df[df["_ts"].dt.month == m]
        label = f"{calendar.month_abbr[m]}-{str(year)[2:]}"
        rows.append({
            "Mes": label,
            "Online":  sub[sub["Fuente"] == "Online"]["Inversión (MXN)"].sum(),
            "Offline": sub[sub["Fuente"] == "Offline"]["Inversión (MXN)"].sum(),
            "OOH":     sub[sub["Fuente"] == "OOH"]["Inversión (MXN)"].sum(),
        })
    result = pd.DataFrame(rows)

    # Para 2025 usar datos hardcoded donde el df no tiene datos
    if year == 2025:
        for i, row in result.iterrows():
            m_num = i + 1
            if row[["Online","Offline","OOH"]].sum() == 0 and m_num in HIST_2025:
                on, off, ooh = HIST_2025[m_num]
                result.at[i, "Online"]  = on
                result.at[i, "Offline"] = off
                result.at[i, "OOH"]     = ooh

    return result


def make_stacked_bar(chart_df: pd.DataFrame, title: str = "") -> go.Figure:
    colors = {"Online": "#0077b6", "Offline": "#e65100", "OOH": "#2e7d32"}
    fig = go.Figure()
    for fuente in ["Online", "Offline", "OOH"]:
        fig.add_trace(go.Bar(
            name=fuente,
            x=chart_df["Mes"],
            y=chart_df[fuente],
            marker_color=colors[fuente],
            hovertemplate="<b>%{x}</b><br>" + fuente + ": $%{y:,.0f}<extra></extra>",
        ))
    fig.update_layout(
        barmode="stack",
        title=title,
        xaxis_title="",
        yaxis_title="Inversión (MXN)",
        yaxis_tickformat="$,.0f",
        legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="right", x=1),
        height=380,
        margin=dict(l=20, r=20, t=40, b=20),
        plot_bgcolor="white",
        paper_bgcolor="white",
    )
    fig.update_yaxes(gridcolor="#f0f0f0")
    return fig


def brand_table(df_main: pd.DataFrame, grupos: set, period_curr, period_prev=None) -> pd.DataFrame:
    """Tabla de inversión por marca para una lista de grupos."""
    df = df_main.copy()
    df["_ts"] = pd.to_datetime(df["Año-mes"], errors="coerce")
    df["_group"] = df["#Grupo"].str.upper()
    df = df[df["_group"].isin({g.upper() for g in grupos})].copy()

    if df.empty:
        return pd.DataFrame(columns=["Marca","Inv. Original","Mes Anterior","Mes Actual","▲/▼"])

    def period_inv(p):
        if p is None:
            return pd.Series(dtype=float)
        sub = df[df["_ts"].dt.to_period("M") == p]
        return sub.groupby("_group")["Inversión (MXN)"].sum()

    inv_curr = period_inv(period_curr)
    inv_prev = period_inv(period_prev)
    inv_total = df.groupby("_group")["Inversión original"].sum()

    marcas = sorted(set(inv_curr.index) | set(inv_prev.index) | set(inv_total.index))
    rows = []
    for marca in marcas:
        curr = inv_curr.get(marca, 0)
        prev = inv_prev.get(marca, 0)
        orig = inv_total.get(marca, 0)
        diff = curr - prev
        arrow = f"▲ ${diff:,.0f}" if diff > 0 else (f"▼ ${abs(diff):,.0f}" if diff < 0 else "–")
        rows.append({"Marca": marca, "Inv. Original": f"${orig:,.0f}",
                     "Mes Anterior": f"${prev:,.0f}", "Mes Actual": f"${curr:,.0f}", "▲/▼": arrow})
    return pd.DataFrame(rows)


def metrics_row(df_main: pd.DataFrame):
    """Devuelve dict con totales por grupo para periodo más reciente en df."""
    df = df_main.copy()
    df["_ts"] = pd.to_datetime(df["Año-mes"], errors="coerce")
    total = df["Inversión (MXN)"].sum()
    def grp_sum(grupos):
        return df[df["#Grupo"].str.upper().isin({g.upper() for g in grupos})]["Inversión (MXN)"].sum()
    return {
        "Total": total,
        "GWM":   grp_sum(GWM_GRUPOS),
        "JAC":   grp_sum(JAC_GRUPOS),
        "KAVAK": grp_sum(KAVAK_GRUPOS),
        "BAIC":  grp_sum(BAIC_GRUPOS),
    }

# ─────────────────────────────────────────────────────────────────────────────
# EXPORTACIÓN
# ─────────────────────────────────────────────────────────────────────────────
def export_xlsx(df: pd.DataFrame) -> bytes:
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        df_out = df.copy()
        df_out["Año-mes"] = pd.to_datetime(df_out["Año-mes"], errors="coerce")
        for col in LAYOUT_COLS:
            if col not in df_out.columns:
                df_out[col] = np.nan
        df_out[LAYOUT_COLS].to_excel(writer, sheet_name="Investment Raw data", index=False)
    return buf.getvalue()


def export_csv(df: pd.DataFrame) -> bytes:
    """CSV formato Power BI: Año-mes como 'jan2026, usa Inversión F30."""
    df_out = df.copy()
    df_out["_ts"] = pd.to_datetime(df_out["Año-mes"], errors="coerce")
    df_out["Año-mes"] = df_out["_ts"].apply(
        lambda t: f"'{MES_ABBR_EN[t.month]}{t.year}" if pd.notna(t) else ""
    )
    df_out["Inversión (MXN)"] = df_out["Inversión F30"]
    csv_cols = ["Fuente","Año","Año-mes","Inversión (MXN)","#Grupo","Categoría","Producto","medio","medio2","modelo"]
    for col in csv_cols:
        if col not in df_out.columns:
            df_out[col] = np.nan
    return df_out[csv_cols].to_csv(index=False, encoding="utf-8-sig").encode("utf-8-sig")

# ─────────────────────────────────────────────────────────────────────────────
# INTERFAZ PRINCIPAL
# ─────────────────────────────────────────────────────────────────────────────
def main():
    # ── Título ────────────────────────────────────────────────────────────────
    st.markdown("""
    <h1 style='margin-bottom:0'>Inversión Automotriz</h1>
    <p style='color:#888;font-size:.95rem;margin-top:4px'>
      Inversión de la industria automotriz &nbsp;·&nbsp; Online &nbsp;·&nbsp; Offline &nbsp;·&nbsp; OOH
    </p>
    """, unsafe_allow_html=True)

    # ── Sidebar ────────────────────────────────────────────────────────────────
    with st.sidebar:
        st.markdown("### Fuentes activas")
        st.markdown("""
        <span class='pill pill-online'>Online</span>
        <span class='pill pill-radio'>Radio</span>
        <span class='pill pill-tele'>Tele</span>
        <span class='pill pill-magazine'>Magazine</span>
        <span class='pill pill-newspaper'>Newspaper</span>
        """, unsafe_allow_html=True)
        st.divider()

        # ── Multiplicadores ────────────────────────────────────────────────
        st.markdown("### Multiplicadores")
        if "mult" not in st.session_state:
            st.session_state["mult"] = DEFAULT_MULT.copy()
        with st.expander("Ver / editar"):
            mult = st.session_state["mult"]
            r1 = st.number_input("Radio M1",        value=mult["radio"][0],        step=0.01, format="%.2f")
            r2 = st.number_input("Radio M2",        value=mult["radio"][1],        step=0.01, format="%.2f")
            ta1= st.number_input("TV Abierta M1",   value=mult["tele_abierta"][0], step=0.01, format="%.2f")
            ta2= st.number_input("TV Abierta M2",   value=mult["tele_abierta"][1], step=0.01, format="%.2f")
            tp1= st.number_input("TV Paga/Local M1",value=mult["tele_paga"][0],    step=0.01, format="%.2f")
            tp2= st.number_input("TV Paga/Local M2",value=mult["tele_paga"][1],    step=0.01, format="%.2f")
            om = st.number_input("OOH MXN factor",  value=mult["ooh_mxn_factor"],  step=0.1,  format="%.1f")
            of = st.number_input("OOH F30 factor",  value=mult["ooh_f30_factor"],  step=0.1,  format="%.2f")
            if st.button("Guardar multiplicadores"):
                st.session_state["mult"] = {
                    "radio": (r1, r2), "tele_abierta": (ta1, ta2),
                    "tele_paga": (tp1, tp2), "online": (1.0, 1.30),
                    "ooh_mxn_factor": om, "ooh_f30_factor": of,
                }
                st.success("Guardado")

    mult = st.session_state.get("mult", DEFAULT_MULT)

    # ── Upload automotriz.xlsx ─────────────────────────────────────────────────
    st.markdown("---")
    st.markdown("### 1. Sube automotriz.xlsx")
    automotriz_file = st.file_uploader(
        "automotriz.xlsx",
        type=["xlsx"],
        key="automotriz_file",
        label_visibility="collapsed",
    )

    if not automotriz_file:
        st.caption("Sube automotriz.xlsx para continuar.")
        st.stop()

    automotriz_bytes = automotriz_file.read()

    # ── Upload archivos fuente ─────────────────────────────────────────────────
    st.markdown("### 2. Sube archivos de inversión")

    tab_gwm, tab_jac, tab_kvk = st.tabs(["GWM", "JAC", "KAVAK · BAIC · Industria"])
    with tab_gwm:
        files_gwm = st.file_uploader(
            "Archivos GWM",
            type=["xlsx","xls"], accept_multiple_files=True, key="gwm_files",
            label_visibility="collapsed",
        )
    with tab_jac:
        files_jac = st.file_uploader(
            "Archivos JAC",
            type=["xlsx","xls"], accept_multiple_files=True, key="jac_files",
            label_visibility="collapsed",
        )
    with tab_kvk:
        files_kvk = st.file_uploader(
            "Archivos KAVAK / BAIC / Industria",
            type=["xlsx","xls"], accept_multiple_files=True, key="kvk_files",
            label_visibility="collapsed",
        )

    all_files = (files_gwm or []) + (files_jac or []) + (files_kvk or [])
    if not all_files:
        st.caption("Sube al menos un archivo de inversión para continuar.")
        st.stop()

    # ── Mes / Año ───────────────────────────────────────────────────────────────
    st.markdown("### 3. Selecciona periodo")
    col_m, col_y, _ = st.columns([2, 2, 6])
    with col_m:
        mes_num = st.selectbox(
            "Mes", list(range(1, 13)),
            format_func=lambda m: calendar.month_name[m],
            index=datetime.now().month - 1,
        )
    with col_y:
        anio = st.selectbox("Año", [2025, 2026, 2027], index=1)

    periods = {pd.Period(f"{anio}-{mes_num:02d}", freq="M")}
    period_curr = next(iter(periods))
    period_prev = (
        pd.Period(f"{anio}-{mes_num-1:02d}", freq="M") if mes_num > 1
        else pd.Period(f"{anio-1}-12", freq="M")
    )

    # ── Botones ────────────────────────────────────────────────────────────────
    st.markdown("---")
    col_btn, col_re = st.columns([3, 1])
    with col_btn:
        run_btn = st.button("Procesar archivos", type="primary", use_container_width=True)
    with col_re:
        re_run = st.button("Volver a correr", use_container_width=True)

    if re_run:
        for key in ["results_df", "df_combined", "automotriz_updated"]:
            st.session_state.pop(key, None)
        st.rerun()

    if run_btn or st.session_state.get("results_df") is not None:
        if run_btn:
            # Procesar
            show_car_animation()
            prog = st.progress(0, text="Procesando GWM...")
            df_gwm = process_gwm(files_gwm or [], periods, mult)
            prog.progress(33, text="Procesando JAC...")
            df_jac = process_jac(files_jac or [], periods, mult)
            prog.progress(66, text="Procesando KAVAK / BAIC...")
            df_kvk = process_kavak_baic_industria(files_kvk or [], periods, mult)
            prog.progress(100, text="Listo")
            prog.empty()

            df_all = pd.concat([df_gwm, df_jac, df_kvk], ignore_index=True)
            if df_all.empty:
                st.warning("No se encontraron datos para el periodo seleccionado.")
                st.stop()

            st.session_state["results_df"] = df_all

            # Combinar con automotriz en memoria
            try:
                df_combined, rb, ra = merge_with_automotriz(df_all, automotriz_bytes)
                st.session_state["df_combined"]       = df_combined
                st.session_state["rows_added"]        = ra - rb
                st.session_state["automotriz_updated"] = build_automotriz_xlsx(df_combined, automotriz_bytes)
            except Exception as e:
                st.session_state["df_combined"]  = None
                st.session_state["merge_error"]  = str(e)

        # ── Resultados ────────────────────────────────────────────────────────
        df_all = st.session_state["results_df"]

        # Conteo por fuente
        st.markdown("---")
        counts = df_all["Fuente"].value_counts()
        total_rows = len(df_all)
        added = st.session_state.get("rows_added", total_rows)

        col1, col2, col3, col4 = st.columns(4)
        col1.metric("Registros nuevos", f"{total_rows:,}")
        col2.metric("Online",  f"{counts.get('Online',0):,}")
        col3.metric("Offline", f"{counts.get('Offline',0):,}")
        col4.metric("OOH",     f"{counts.get('OOH',0):,}")

        if st.session_state.get("automotriz_updated"):
            ts_label = f"{calendar.month_abbr[mes_num]}{anio}"
            st.success(f"{added:,} filas nuevas incluidas en el archivo.")
            st.download_button(
                "Descargar automotriz actualizado",
                data=st.session_state["automotriz_updated"],
                file_name=f"automotriz_{ts_label}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                type="primary",
            )
        elif "merge_error" in st.session_state:
            st.error(f"Error al combinar: {st.session_state['merge_error']}")

        # df para análisis: usar el combinado si existe, si no solo los nuevos
        df_main = st.session_state.get("df_combined")
        if df_main is None:
            df_main = df_all.copy()

        df_main["Año-mes"] = pd.to_datetime(df_main["Año-mes"], errors="coerce")

        # ── Métricas ──────────────────────────────────────────────────────────
        st.markdown("---")
        st.markdown("#### Inversión total acumulada en automotriz.xlsx")
        met = metrics_row(df_main)
        cols = st.columns(5)
        for col_w, (k, v) in zip(cols, met.items()):
            col_w.markdown(
                f"<div class='metric-card'><div class='metric-label'>{k}</div>"
                f"<div class='metric-value'>${v/1e6:.1f}M</div></div>",
                unsafe_allow_html=True,
            )

        # ── Gráfica por año ───────────────────────────────────────────────────
        st.markdown("---")
        chart_year = st.session_state.get("chart_year", 2026)
        c1, c2, _ = st.columns([1, 1, 8])
        with c1:
            if st.button("2026", type="primary" if chart_year == 2026 else "secondary"):
                st.session_state["chart_year"] = 2026
                st.rerun()
        with c2:
            if st.button("2025", type="primary" if chart_year == 2025 else "secondary"):
                st.session_state["chart_year"] = 2025
                st.rerun()

        chart_df = build_chart_data(df_main, chart_year)
        st.plotly_chart(
            make_stacked_bar(chart_df, f"Inversión por fuente – {chart_year}"),
            use_container_width=True,
        )

        # ── Tabs por grupo ────────────────────────────────────────────────────
        st.markdown("---")
        tab_total, tab_gwm_r, tab_jac_r, tab_kvk_r, tab_baic_r = st.tabs(
            ["Total", "GWM", "JAC", "KAVAK", "BAIC"]
        )

        all_grupos = GWM_GRUPOS | JAC_GRUPOS | KAVAK_GRUPOS | BAIC_GRUPOS

        def render_tab(grupos, label):
            bt = brand_table(df_main, grupos, period_curr, period_prev)
            st.dataframe(bt, use_container_width=True, hide_index=True)
            sub = df_main[df_main["#Grupo"].str.upper().isin({g.upper() for g in grupos})].copy()
            cdf = build_chart_data(sub, chart_year)
            if cdf[["Online","Offline","OOH"]].sum().sum() > 0:
                st.plotly_chart(
                    make_stacked_bar(cdf, f"{label} – {chart_year}"),
                    use_container_width=True,
                )

        with tab_total:
            render_tab(all_grupos, "Total industria")
        with tab_gwm_r:
            render_tab(GWM_GRUPOS, "GWM")
        with tab_jac_r:
            render_tab(JAC_GRUPOS, "JAC")
        with tab_kvk_r:
            render_tab(KAVAK_GRUPOS, "KAVAK")
        with tab_baic_r:
            render_tab(BAIC_GRUPOS, "BAIC")

        # ── Exportar ──────────────────────────────────────────────────────────
        st.markdown("---")
        st.markdown("#### Exportar datos procesados")
        ecol1, ecol2, _ = st.columns([2, 2, 6])
        ts_label = f"{calendar.month_abbr[mes_num]}{anio}"
        with ecol1:
            st.download_button(
                "Descargar XLSX (solo nuevos)",
                data=export_xlsx(df_all),
                file_name=f"nuevos_{ts_label}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        with ecol2:
            st.download_button(
                "Descargar CSV (Power BI)",
                data=export_csv(df_all),
                file_name=f"layout_{ts_label}.csv",
                mime="text/csv",
            )


if __name__ == "__main__":
    main()
